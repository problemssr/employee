from django.shortcuts import render, HttpResponse, redirect
from api import models
from api.utils.pagenation import Pagenation
from api.utils.BootStrap import BootStrapModelForm
from django import forms
from django.core.validators import RegexValidator, ValidationError
from api.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm


def depart_list(request):
    """ 部门列表 """

    # 在数据库中获取所有部门信息
    querySet = models.Department.objects.all()

    page_object = Pagenation(request, querySet, page_size=5)
    context = {
        'depart': page_object.page_queryset,
        'page_string': page_object.html()
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, 'depart_add.html')
    else:
        # 获取用户post数据
        title = request.POST.get("title")
        # 保存数据库
        models.Department.objects.create(title=title)
        # 重定向回部门列表
        return redirect("/depart/list/")


def depart_delete(request):
    """ 删除部门 """
    # 获取ID
    nid = request.GET.get("nid")
    # 删除
    models.Department.objects.filter(id=nid).delete()
    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_edit(request):
    """ 编辑部门 """
    # 根据nid获取数据
    nid = request.GET.get("nid")
    if request.method == "GET":
        rows = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"row": rows})

    else:
        # 获取用户提交的标题
        title = request.POST.get("title")
        # 根据ID找到数据库中数据并更新
        models.Department.objects.filter(id=nid).update(title=title)

        return redirect("/depart/list")


def depart_mutli(request):
    """批量上传（Excel）"""
    from openpyxl import load_workbook

    # 读取Excel文件内容

    # 1.获取用户上传文件对象
    file_obj = request.FILES.get("Excel")

    # 2.对象传递给openpyxl,由openpyxl读取文件内容
    wb = load_workbook(file_obj.name)  # 加载csv
    sheet = wb.worksheets[0]  # 获取Excel文件对象（第一张表）

    # cell=sheet.cell(1,2)   #cell获取对象
    # print(cell.value)

    # 3.循环获取每行数据
    for row in sheet.iter_rows(min_row=1, max_row=10):
        text = row[1].value
        exist = models.Department.objects.filter(title=text).exists()
        if not exist:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
